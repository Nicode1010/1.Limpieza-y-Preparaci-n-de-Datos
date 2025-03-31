import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Cargar el archivo Excel
def cargar_datos(file_path):
    try:
        df = pd.read_excel(file_path)
        return df
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        exit()

# Función para corregir valores de Facultad y Programa
def corregir_facultad_programa(df):
    # Crear diccionarios de referencia
    facultades_por_programa = df.groupby("Programa")["Facultad"].first().to_dict()
    programas_por_facultad = df.groupby("Facultad")["Programa"].first().to_dict()
    
    # Corregir Facultad si está vacía y el programa es conocido
    df["Facultad"].fillna(df["Programa"].map(facultades_por_programa), inplace=True)
    
    # Corregir Programa si está vacío y la facultad es conocida
    df["Programa"].fillna(df["Facultad"].map(programas_por_facultad), inplace=True)
    
    # Rellenar valores desconocidos
    df["Facultad"].fillna("Desconocido", inplace=True)
    df["Programa"].fillna("Desconocido", inplace=True)
    
    return df

# Función para corregir Año_Ingreso
def corregir_anio_ingreso(df):
    def extraer_anio(valor):
        if isinstance(valor, str) and "-" in valor:
            return min(map(int, valor.split("-")))  # Tomar el menor año
        try:
            return int(valor)  # Convertir a entero si es posible
        except:
            return np.nan  # Dejar como NaN si no es convertible
    
    df["Año_Ingreso"] = df["Año_Ingreso"].apply(extraer_anio)
    return df

# Función para resaltar valores problemáticos en amarillo
def resaltar_problemas(df, output_path):
    df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value in ["Desconocido", "N/A", np.nan] or isinstance(cell.value, str) and "-" in cell.value:
                cell.fill = fill
    
    wb.save(output_path)

# Ejecutar el proceso
file_path = "estudiantes.xlsx"
df = cargar_datos(file_path)
df = corregir_facultad_programa(df)
df = corregir_anio_ingreso(df)
df.drop_duplicates(subset="ID_Estudiante", keep="first", inplace=True)

# Guardar el archivo limpio con resaltado de problemas
output_path = "estudiantes_limpios.xlsx"
resaltar_problemas(df, output_path)
print(f"✅ Archivo limpio guardado en: {output_path}")
